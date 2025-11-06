"""
PII Detection and Anonymization Framework

This script generates synthetic datasets, annotates PII entities, trains a custom spaCy NER model,
evaluates its performance, and anonymizes detected PII from text data.

"""

"""Importing all required Libraries"""
import random
import re
import ast
import warnings

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

from faker import Faker

import spacy
from spacy.training import Example, offsets_to_biluo_tags
from spacy.util import minibatch, compounding

from sklearn.metrics import (
    precision_score,
    recall_score,
    f1_score,
    accuracy_score,
    precision_recall_fscore_support,
    confusion_matrix,
    precision_recall_curve,
    roc_curve,
    roc_auc_score,
)

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

"""Generation of Training Dataset"""

# Initialize Faker
fake = Faker()

# Function to generate phone numbers in a specific format
def generate_phone_number():
    formats = [
        '+91 ##########',
        '+## ##########',
        '+### ##########'
    ]
    format_choice = fake.random.choice(formats)
    return fake.numerify(format_choice)

# Function to generate synthetic PII data
def generate_pii_data(num_samples):
    data = {
        "name": [fake.name() for _ in range(num_samples)],
        "credit_card": [fake.credit_card_full() for _ in range(num_samples)],
        "email": [fake.email() for _ in range(num_samples)],
        "url": [fake.url() for _ in range(num_samples)],
        "phone": [generate_phone_number() for _ in range(num_samples)],
        "address": [fake.address() for _ in range(num_samples)],
        "company": [fake.company() for _ in range(num_samples)],
        "ssn": [fake.ssn() for _ in range(num_samples)]
    }
    return pd.DataFrame(data)

# Function to randomly remove full stops
def remove_random_full_stops(text, removal_probability=0.3):
    if random.random() < removal_probability:
        text = text.replace('.', '', random.randint(1, text.count('.')))
    return text

# Generate Dataset with any Number of Samples
pii_dataset = generate_pii_data(45000)

# Combined sentence/audit templates with PII data
sentence_templates = [

('''To The Members of {company},
Report on the audit of the Standalone Financial Statements
Key Audit Matters Auditors’ response to Key Audit Matters
Property, Plant & Equipment and Intangible Assets
There are areas where management judgement impacts the carrying value of property, plant and equipment, intangible assets and their respective depreciation/amortisation rates. These include the decision to capitalize or expense costs; the annual asset life review; the timeliness of the capitalization of assets and the use of management assumptions and estimates for the determination or the measurement and recognition criteria for assets retired from active use. Due to the materiality in the context of the Balance Sheet of the Company and the level of judgement and estimates required, we consider this to be an area of significance.
We assessed the controls in place over the fixed asset cycle, evaluated the appropriateness of the capitalization process, performed tests of details on costs capitalized, the timeliness of the capitalization of the assets and the de-recognition criteria for assets retired from active use. In performing these procedures, we reviewed the judgements made by management including the nature of underlying costs capitalized; determination of realizable value of the assets retired from active use; the appropriateness of asset lives applied in the calculation of depreciation/amortization; the useful lives of assets prescribed in Schedule II to the Act and the useful lives of certain assets as per the technical assessment of the management. We observed that the management has regularly reviewed the aforesaid judgements and there are no material changes.
Opinion
We have audited the accompanying standalone financial statements of {company}, which comprise the Balance Sheet as at March 31, 2023, the Statement of Profit and Loss (including Other Comprehensive Income), the Statement of Changes in Equity and the Statement of Cash Flows for the year then ended and notes to the standalone financial statements including a summary of significant accounting policies and other explanatory information in which are incorporated the financial statements for the year ended on that date audited by the Branch Auditors of the Company’s one Branch, namely Research & Development (R&D) division situated at {address}. In our opinion and to the best of our information and according to the explanations given to us, the aforesaid standalone financial statements give the information required by the Companies Act, 2013 (the “Act”) in the manner so required and give a true and fair view in conformity with Indian Accounting Standards specified under section 133 of the Act read with the Companies (Indian Accounting Standards) Rules 2015, as amended and other accounting principles generally accepted in India, of the state of affairs of the Company as at March 31, 2023, and total comprehensive income (comprising of profit and other comprehensive income), changes in equity and its cash flows for the year ended on that date.
Basis for opinion
We conducted our audit of the standalone financial statements in accordance with the Standards on Auditing (SAs) specified under section 143(10) of the Act. Our responsibilities under those Standards are further described in the Auditors’ Responsibilities for the Audit of the standalone financial statements section of our report. We are independent of the Company in accordance with the Code of Ethics issued by the Institute of Chartered Accountants of India (“ICAI”), together with the ethical requirements that are relevant to our audit of the standalone financial statements under the provisions of the Act and "the Rules" thereunder, and we have fulfilled our other ethical responsibilities in accordance with these requirements and the ICAI’s Code of Ethics. We believe that the audit evidence we have obtained is sufficient and appropriate to provide a basis for our opinion on the standalone financial statements.
Key audit matters
Key audit matters are those matters that, in our professional judgment, were of most significance in our audit of the standalone financial statements of the current period. These matters were addressed in the context of our audit of the standalone financial statements as a whole, and in forming our opinion thereon, and we do not provide a separate opinion on these matters. We have determined the matters described below to be the key audit matters to be communicated in our report.'''),

('''Following the annual compliance schedule, a thorough evaluation of the internal controls within {company} was conducted. This document outlines the methodologies adopted, key findings, and recommendations for strengthening internal processes and ensuring regulatory compliance. The objective of this review was to assess the effectiveness of the existing controls and identify potential areas for enhancement.
The review process included both quantitative and qualitative assessments, leveraging advanced data analytics tools and direct observation techniques. The methodology encompassed:
Risk Assessment:
Identification and prioritization of risk areas within the financial and operational domains
Deployment of risk management frameworks to evaluate the potential impact and likelihood of identified risks
Control Testing:
Execution of control tests to verify the functionality and effectiveness of control measures
Analysis of control gaps and deficiencies in critical areas
Stakeholder Interviews:
Conducting interviews with key personnel to gather insights on control environments and operational challenges
Evaluation of the awareness and understanding of control policies among staff
Several critical findings emerged from the review, indicating areas that require immediate attention and corrective actions:
Financial Control Deficiencies:
Inconsistent application of accounting policies leading to discrepancies in financial reporting
Lack of adequate documentation for significant financial transactions
Operational Control Gaps:
Inefficiencies in the procurement process resulting in unauthorized purchases
Insufficient monitoring of inventory levels causing stock variances
Compliance Shortcomings:
Non-compliance with internal audit recommendations from previous assessments
Delays in regulatory filings and updates
To address the identified deficiencies and enhance the control environment, the following actions are recommended:
Financial Controls:
Standardization of accounting procedures across all departments
Complementation of a centralized documentation system for financial transactions
Operational Controls:
Revision of the procurement policy to include stricter approval processes
Regular inventory audits to ensure accuracy and accountability
Compliance Enhancements:
Establishment of a compliance oversight committee to monitor adherence to audit recommendations
Timely updating and submission of regulatory documents
The review highlights the necessity for continuous improvement in internal controls to mitigate risks and ensure compliance. The implementation of the recommended actions will significantly enhance the operational efficiency and financial integrity of {company}
Attached to this document are detailed reports and evidence supporting the findings and recommendations. The Internal Compliance Unit is available for further discussions and clarifications
For any questions or clarifications regarding this document, please contact the Internal Compliance Unit at {email} or {phone}. Physical correspondence can be directed to {address}. Please also include the last four digits of your SSN: {ssn}.
This document contains proprietary information of {company}. Unauthorized use or disclosure of the contents is strictly prohibited. All related communications should be directed to authorized personnel only.
Please acknowledge receipt of this document by providing the last four digits of your {credit_card}, your full name {name}, and the associated {url} to our secure email. This step is crucial for maintaining the security and confidentiality of our internal review process.'''),

('''Independent Auditor's Report on the Standalone Financial Statements of {company} for the fiscal year ended March 31, 2023. The audit was conducted in accordance with the Standards on Auditing specified under section 143(10) of the Companies Act, 2013. Our audit involved performing procedures to obtain audit evidence about the amounts and disclosures in the standalone financial statements. The procedures selected depend on the auditor's judgment, including the assessment of the risks of material misstatement of the financial statements, whether due to fraud or error. In making those risk assessments, the auditor considers internal control relevant to the company's preparation and fair presentation of the standalone financial statements in order to design audit procedures that are appropriate in the circumstances. Our audit also included evaluating the appropriateness of accounting policies used and the reasonableness of accounting estimates made by management, as well as evaluating the overall presentation of the standalone financial statements. We believe that the audit evidence we have obtained is sufficient and appropriate to provide a basis for our audit opinion. Our opinion, based on our audit, is that the accompanying standalone financial statements give a true and fair view of the financial position of the company as of March 31, 2023, and of its financial performance and its cash flows for the year then ended in accordance with the Indian Accounting Standards prescribed under section 133 of the Act read with the Companies (Indian Accounting Standards) Rules, 2015, as amended. Key audit matters are those matters that, in our professional judgment, were of most significance in our audit of the standalone financial statements of the current period. These matters were addressed in the context of our audit of the standalone financial statements as a whole, and in forming our opinion thereon, and we do not provide a separate opinion on these matters. The management and Board of Directors of {company} are responsible for the matters stated in section 134(5) of the Companies Act, 2013 with respect to the preparation of these standalone financial statements that give a true and fair view of the financial position, financial performance, and cash flows of the Company in accordance with the Indian Accounting Standards (Ind AS) and other accounting principles generally accepted in India. This responsibility also includes maintenance of adequate accounting records in accordance with the provisions of the Act for safeguarding the assets of the Company and for preventing and detecting frauds and other irregularities; selection and application of appropriate accounting policies; making judgments and estimates that are reasonable and prudent; and design, implementation and maintenance of adequate internal financial controls that were operating effectively for ensuring the accuracy and completeness of the accounting records, relevant to the preparation and presentation of the standalone financial statements that give a true and fair view and are free from material misstatement, whether due to fraud or error. In preparing the standalone financial statements, management is responsible for assessing the Company’s ability to continue as a going concern, disclosing, as applicable, matters related to going concern and using the going concern basis of accounting unless management either intends to liquidate the Company or to cease operations, or has no realistic alternative but to do so. The Board of Directors are also responsible for overseeing the Company’s financial reporting process. For any queries, please contact {name} at {address} or {phone}. Additional information can be found at {url}. The last four digits of your SSN {ssn} may be requested for verification purposes during any queries.'''),

('''We have conducted a thorough review of the tax compliance practices followed by {company} for the fiscal year ending March 31, 2023. Our examination included a detailed analysis of corporate tax returns, GST filings, and withholding tax submissions across all divisions. The review focused on ensuring compliance with the latest amendments in tax laws and regulations.

Corporate Tax Overview
The corporate tax computation for {company} was cross-verified against the financial statements audited by our internal team. The tax liability was calculated considering various deductions under section 80C, 80D, and other relevant sections of the Income Tax Act. The total taxable income stood at INR 500 Crores, with an effective tax rate of 25%.

Key points include:

Depreciation Deductions: Claimed as per the Income Tax Act, aligned with the rates prescribed under Schedule II. The assets located at {address} were correctly depreciated using the Written Down Value (WDV) method. The details of high-value assets have been corroborated with the asset register maintained at the corporate office.

Tax Credits: The company has utilized carry-forward losses from previous financial years to offset the current tax liability, reducing the net payable tax. The adjusted tax liability has been duly filed with the tax authorities.

Deductions: The deductions for contributions to the Employee Provident Fund (EPF) and Gratuity are in compliance with sections 80C and 80D. However, we noted a delay in the deposit of EPF contributions for some employees whose SSNs {ssn} end with ‘4567’. This delay has been flagged, and a provision for potential interest and penalties has been recommended.

GST Compliance
The Goods and Services Tax (GST) compliance was reviewed in detail:

GST Payments: All GST payments were made on time except for minor discrepancies in the month of July. The shortfall of INR 2 Lakhs in GST payments for {company}’s manufacturing unit at {address} was rectified in subsequent months, with interest computed at 18% p.a.

Input Tax Credit (ITC): ITC claims were verified against the purchase invoices. The ITC related to capital goods purchased by the Research & Development (R&D) division were adequately accounted for. However, it was observed that certain invoices, particularly from vendors identified by the URL {url}, were not uploaded on the GST portal within the stipulated time, leading to an ITC reversal.

Reconciliation: A reconciliation of GSTR-3B with GSTR-2A was performed, revealing minor mismatches which have been communicated to the concerned department. The finance team has been instructed to follow up with vendors whose SSNs {ssn} match records ending in '7890' to ensure timely filing.

Withholding Tax (TDS)
Withholding tax (TDS) was analyzed across various payments made during the year:

Salaries: TDS on salaries was deducted as per Section 192 of the Income Tax Act. Employee details, including SSNs {ssn}, were cross-checked with the HR records. A mismatch was found in the TDS calculations for employees whose SSNs end in ‘1234’ due to incorrect consideration of their investment declarations.

Professional Fees: TDS under Section 194J was reviewed, with a specific focus on payments exceeding INR 30,000. One such transaction involving a payment to {name} was identified where TDS was not deducted. The legal team has been notified, and a rectification process has been initiated.

Rent Payments: TDS on rent payments was calculated correctly, but it was observed that rent agreements for premises at {address} lacked proper documentation. The agreements are currently being reviewed to ensure compliance.

Audit Recommendations
Based on our review, we recommend the following actions to mitigate tax risks:

Timely Payment of Taxes: Ensure that all taxes, including GST and TDS, are paid within the due dates to avoid interest and penalties. The finance team should regularly review the payment schedules, particularly for transactions involving large sums.

Documentation: Improve the documentation process, especially for transactions involving high-value assets and payments. Ensure that all contracts and agreements are updated and compliant with tax regulations.

Employee Training: Conduct training sessions for the finance and HR teams on the latest tax amendments and compliance requirements. Emphasize the importance of accurate TDS calculations and timely tax payments.

Automation of Processes: Consider implementing tax compliance software to automate GST reconciliation, TDS computation, and other tax-related processes. This will reduce manual errors and ensure adherence to compliance timelines.

Conclusion
We have attached a detailed report with the findings and recommendations. The finance team at {company} should review this report and initiate the necessary actions. For any queries or further clarifications, please contact {name} at {email} or {phone}. All physical correspondence can be directed to our office at {address}.

This review reflects our commitment to ensuring that {company} remains compliant with all tax regulations. We appreciate your cooperation during the audit process and look forward to your prompt action on the recommendations.

'''),

('''To Whom It May Concern,

This letter is to confirm that {name}, holding Social Security Number (SSN) {ssn}, residing at {address}, has filed their tax returns for the fiscal year ending March 31, 2023. The tax filings have been processed under the IRS Tax Identification Number (TIN) associated with the company {company}.

The individual’s total income for the fiscal year amounted to $125,000, including salaries, bonuses, and other forms of income. The detailed breakdown of the income sources is as follows:

1. Salary from {company}: $100,000
2. Bonus and Incentives: $15,000
3. Other Income (Investments, Dividends, etc.): $10,000

The total federal tax liability for the year is $25,000, which has been fully paid by the taxpayer. The tax payments were made using the credit card ending in {credit_card} and were processed through the IRS online payment portal. Please note that this tax statement is generated in accordance with the income tax laws applicable in the United States.

For any queries or further clarifications, you may contact {name} at {email} or {phone}. Additional documents supporting the tax filings can be requested by visiting {url}.

This document is confidential and should be handled in accordance with data protection regulations to prevent unauthorized access to sensitive information.
'''),

('''Subject: Annual Tax Filing Confirmation for Fiscal Year 2023

Dear {name},

We are pleased to inform you that your tax return for the fiscal year ending March 31, 2023, has been successfully filed and processed by {company}. The filing was completed using your Social Security Number (SSN) {ssn}, and the confirmation number is associated with the TIN {ssn} registered under {company}.

Your gross income for the year was reported as $150,000, which includes:

- Employment Income from {company}: $120,000
- Capital Gains: $20,000
- Interest and Dividends: $10,000

The total tax due for the fiscal year was calculated at $30,000. This amount has been paid in full through a payment transaction completed on March 28, 2024, using the credit card ending in {credit_card}. Your tax records indicate that you are eligible for a tax refund of $2,000, which will be credited to your bank account on file.

Please ensure that all records related to this tax filing, including the payment receipt, are stored securely. Should you require any further assistance or have questions about your tax return, please do not hesitate to contact our customer service department at {email} or by calling {phone}. You may also visit our website {url} for more information.

This statement is intended for the use of {name} and contains sensitive information that must be kept confidential. Any unauthorized use, dissemination, or copying of this document is strictly prohibited.
'''),

('''To: {name}
SSN: {ssn}
TIN: {ssn}
Address: {address}

Subject: Confirmation of Tax Filing for FY 2023

Dear {name},

Your tax filing for the fiscal year ending March 31, 2023, has been successfully processed. The filing was conducted using your TIN {ssn} registered under the IRS. Your total income for the fiscal year was reported as $175,000, comprising the following sources:

1. Salary from {company}: $140,000
2. Investment Income: $25,000
3. Other Earnings: $10,000

The total tax payable for this fiscal year was calculated at $35,000. The payment was made on March 25, 2024, using the credit card associated with the number ending in {credit_card}. Please retain this statement as proof of payment and tax compliance.

Should you need to amend any details or have inquiries, you can contact us at {email} or by phone at {phone}. Additional information and related services can be accessed through {url}.

Please note that this document contains confidential information, and unauthorized access or distribution is strictly prohibited. Keep this document in a secure location.
'''),

"During the audit of the financial statements for {company}, it was observed that Mr. {name}, the Chief Financial Officer, approved the purchase of assets worth $500,000 on {address}. The payment was processed through the credit card ending in {credit_card}. Mr. {name}'s Social Security Number (SSN) is {ssn}. For any clarifications, please reach out to Mr. {name} at {email} or contact him directly at {phone}. Further details can be accessed at {url}.",

"Customer {name} from {company}, located at {address}, has submitted feedback regarding the recent transaction involving their credit card {credit_card}. They can be reached at {email} or {phone} for further discussions. The customer's Social Security Number (SSN) on file is {ssn}. The feedback was originally submitted through our website at {url}.",

"This is a confirmation that the payment for the invoice number INV-409876535422 from {company} has been successfully processed. The payment was made using the credit card ending in {credit_card} by {name}. The billing address on file is {address}. The Social Security Number (SSN) for {name} is {ssn}. Should you have any inquiries, you may contact {name} via email at {email} or call {phone}. For more information, visit {url}",

"We regret to inform you that a security breach was detected on {company}'s systems, which may have exposed your personal information, including your name ({name}), email ({email}), phone number ({phone}), and Social Security Number (SSN) ({ssn}). The breach was traced back to unauthorized access from IP address 192.0.45. If you notice any suspicious activity on your credit card ending in {credit_card}, please contact us immediately. You can also check for updates on our security measures at {url}. The compromised data was stored at our facility located at {address}.",

"Dear {name}, thank you for creating a new account with {company}. Your registered email is {email}, your contact number is {phone}, and your Social Security Number (SSN) is {ssn}. The account was set up using the billing address {address}, and the primary credit card linked to the account ends in {credit_card}. Please visit {url} to verify your account and update any personal details. If you need assistance, contact our support team",

"This Service Contract between {company} and {name} was entered at {address}. The contract stipulates that all payments will be processed through the credit card provided by {name}, ending in {credit_card}. {name}'s Social Security Number (SSN) is {ssn}. For further reference, correspondence will be sent to {email}, and all communications will be conducted via {phone}. The full contract details are available online at {url}.",

"Dear {name}, we have received your loan application at {company}, and it is currently under review. Your application, submitted on 23/09/2008, includes personal details such as your home address ({address}), email ({email}), contact number ({phone}), and Social Security Number (SSN) ({ssn}). The loan amount requested will be credited to your account associated with the credit card ending in {credit_card}. Please check {url} for real-time updates on your application status.",

"Insurance claim #CLM9076109877 has been initiated by {name} for {company}. The claim, associated with the address {address}, will be processed through the credit card ending in {credit_card}. Our claims department may reach out to you at {phone} or via email at {email} for additional information. {name}'s Social Security Number (SSN) is {ssn}. Claim details are available online at {url}.",

"We are pleased to welcome {name} to {company}. As part of the onboarding process, we have registered your personal details, including your residential address ({address}), contact number ({phone}), email ({email}), and Social Security Number (SSN) ({ssn}). Your corporate credit card, ending in {credit_card}, will be issued within the next five business days. For company policies and other relevant information, please visit {url}.",

"Dear {name}, your subscription with {company} is up for renewal. The subscription associated with the email {email}, phone number {phone}, and billing address {address} will automatically renew using your credit card ending in {credit_card}. Your Social Security Number (SSN) on file is {ssn}. To manage your subscription or for more details, please visit {url}. If you need to update your payment information, contact our support team."

"During the routine financial audit for {company}, it was discovered that Mr. {name}, serving as the CFO, authorized the acquisition of assets totaling $500,000. The transaction occurred at {address} and was paid for using the credit card ending in {credit_card}. Mr. {name} is identified by the Social Security Number (SSN) {ssn}. For additional details, Mr. {name} can be contacted at {email} or via phone at {phone}. Further documentation can be found on our website at {url}",

"Customer feedback has been received from {name} representing {company}, located at {address}. The feedback pertains to a transaction processed through their credit card ending in {credit_card}. You can reach out to {name} for more information at {email} or {phone}. The SSN associated with the customer's profile is {ssn}. The feedback submission was completed through our online portal at {url}.",

"We hereby confirm that the payment for invoice INV-12345 issued by {company} has been successfully completed. The payment was facilitated by {name} using the credit card ending in {credit_card}. The billing address linked to the payment is {address}. The associated SSN for {name} is {ssn}. For inquiries, please contact {name} at {email} or by phone at {phone}. Visit {url} for further details",

"A security incident has been detected on the systems of {company}, potentially compromising your personal data, including your name ({name}), email ({email}), phone number ({phone}), and SSN ({ssn}). The breach was linked to an unauthorized access attempt traced to IP address. If you observe any suspicious activity on your credit card ending in {credit_card}, notify us immediately. Updates on the situation will be posted at {url}. The compromised data was stored at our facility located at {address}",

"Dear {name}, we are pleased to confirm the creation of your new account with {company}. Your registered email address is {email}, and your contact number is {phone}. The account was set up using your billing address {address} and is linked to a credit card ending in {credit_card}. Your Social Security Number (SSN) is {ssn}. Please visit {url} to verify your account and update any personal information. If you require assistance, our support team is available to help.",

"This agreement between {company} and {name} was formalized at {address}. Under the terms of the contract, all payments will be processed via the credit card provided by {name}, ending in {credit_card}. {name}'s Social Security Number (SSN) is {ssn}. All correspondence will be directed to {email}, and further communication can be made via {phone}. Full contract details are accessible online at {url}",

"Dear {name}, your recent loan application at {company} is currently under review. The application, submitted on 09/12/2020, includes your home address ({address}), email ({email}), phone number ({phone}), and SSN ({ssn}). The requested loan amount will be credited to the account linked to the credit card ending in {credit_card}. Check {url} for real-time updates on your application status",

"Insurance claim #CLM0098 has been initiated by {name} with {company}. The claim, related to the address {address}, will be processed using the credit card ending in {credit_card}. Our claims department may contact you at {phone} or {email} for further details. The SSN associated with {name} is {ssn}. Claim details can be accessed at {url}.",

"We are excited to welcome {name} to {company}. As part of your onboarding, we have registered your personal details, including your home address ({address}), contact number ({phone}), email ({email}), and SSN ({ssn}). Your corporate credit card, ending in {credit_card}, will be issued shortly. For more information on company policies, please visit {url}.",

"Dear {name}, your subscription with {company} is approaching its renewal date. The subscription linked to the email {email}, phone number {phone}, and billing address {address} will automatically renew using the credit card ending in {credit_card}. The SSN on file for this account is {ssn}. To manage your subscription or update payment details, visit {url}. If you need further assistance, please contact our support team.",

"During the audit of the financial statements for {company}, it was observed that Mr. {name}, the Chief Financial Officer, approved the purchase of assets worth $500,000 on {address}. The payment was processed through the credit card ending in {credit_card}. Mr. {name}'s Social Security Number (SSN) is {ssn}. For any clarifications, please reach out to Mr. {name} at {email} or contact him directly at {phone}. Further details can be accessed at {url}.",

"Customer {name} from {company}, located at {address}, has submitted feedback regarding the recent transaction involving their credit card {credit_card}. They can be reached at {email} or {phone} for further discussions. The customer's Social Security Number (SSN) on file is {ssn}. The feedback was originally submitted through our website at {url}.",

"This is a confirmation that the payment for the invoice number INV-3066 from {company} has been successfully processed. The payment was made using the credit card ending in {credit_card} by {name}. The billing address on file is {address}. The Social Security Number (SSN) for {name} is {ssn}. Should you have any inquiries, you may contact {name} via email at {email} or call {phone}. For more information, visit {url}.",

"We regret to inform you that a security breach was detected on {company}'s systems, which may have exposed your personal information, including your name ({name}), email ({email}), phone number ({phone}), and Social Security Number (SSN) ({ssn}). The breach was traced back to unauthorized access from IP address 190.22.99. If you notice any suspicious activity on your credit card ending in {credit_card}, please contact us immediately. You can also check for updates on our security measures at {url}. The compromised data was stored at our facility located at {address}.",

"Dear {name}, thank you for creating a new account with {company}. Your registered email is {email}, your contact number is {phone}, and your Social Security Number (SSN) is {ssn}. The account was set up using the billing address {address}, and the primary credit card linked to the account ends in {credit_card}. Please visit {url} to verify your account and update any personal details. If you need assistance, contact our support team.",

"This Service Contract between {company} and {name} was entered into 11/11/2005 at {address}. The contract stipulates that all payments will be processed through the credit card provided by {name}, ending in {credit_card}. {name}'s Social Security Number (SSN) is {ssn}. For further reference, correspondence will be sent to {email}, and all communications will be conducted via {phone}. The full contract details are available online at {url}.",

"Dear {name}, we have received your loan application at {company}, and it is currently under review. Your application, submitted on 1998, includes personal details such as your home address ({address}), email ({email}), contact number ({phone}), and Social Security Number (SSN) ({ssn}). The loan amount requested will be credited to your account associated with the credit card ending in {credit_card}. Please check {url} for real-time updates on your application status.",

"Insurance claim #CLM12345678 has been initiated by {name} for {company}. The claim, associated with the address {address}, will be processed through the credit card ending in {credit_card}. Our claims department may reach out to you at {phone} or via email at {email} for additional information. {name}'s Social Security Number (SSN) is {ssn}. Claim details are available online at {url}.",

"We are pleased to welcome {name} to {company}. As part of the onboarding process, we have registered your personal details, including your residential address ({address}), contact number ({phone}), email ({email}), and Social Security Number (SSN) ({ssn}). Your corporate credit card, ending in {credit_card}, will be issued within the next five business days. For company policies and other relevant information, please visit {url}.",

"Dear {name}, your subscription with {company} is up for renewal. The subscription associated with the email {email}, phone number {phone}, and billing address {address} will automatically renew using your credit card ending in {credit_card}. Your Social Security Number (SSN) on file is {ssn}. To manage your subscription or for more details, please visit {url}. If you need to update your payment information, contact our support team."

]

# Apply the templates to generate sentences with PII data
pii_dataset['text'] = pii_dataset.apply(lambda row: sentence_templates[row.name % len(sentence_templates)].format(
    name=row['name'],
    company=row['company'],
    email=row['email'],
    url=row['url'],
    phone=row['phone'],
    address=row['address'],
    credit_card=row['credit_card'],
    ssn=row['ssn']
), axis=1)

# Apply the random full stop removal
pii_dataset['text'] = pii_dataset['text'].apply(remove_random_full_stops)

# Save the dataset to a CSV file
csv_file_path = r'Training_Set.csv'
pii_dataset.to_csv(csv_file_path, index=False)

# Display the first few rows of the dataset
print(pii_dataset.head())
print(f"Data successfully written to {csv_file_path}")

"""# Annotation of True PII Data Position in Training Dataset"""

# Load the dataset
csv_file_path = r'Training_Set.csv'
pii_dataset = pd.read_csv(csv_file_path)

# Function to annotate PII data in text
def annotate_pii(text, pii_dict):
    annotations = []
    for pii_type, pii_value in pii_dict.items():
        # Escape special characters in PII data for regex
        escaped_pii_value = re.escape(pii_value)
        # Find all matches of PII data in the text
        matches = list(re.finditer(escaped_pii_value, text))
        for match in matches:
            start, end = match.span()
            annotations.append((start, end, pii_type))
    return annotations

# Apply the annotation function to each row
pii_dataset['True Predictions'] = pii_dataset.apply(lambda row: annotate_pii(
    row['text'], {
        'name': row['name'],
        'credit_card': row['credit_card'],
        'email': row['email'],
        'url': row['url'],
        'phone': row['phone'],
        'address': row['address'],
        'company': row['company'],
        'ssn': row['ssn']
    }), axis=1)

# Save the annotated dataset to a new CSV file
annotated_csv_file_path = r'Training_Set.csv'
pii_dataset.to_csv(annotated_csv_file_path, index=False)

# Display the first few rows of the annotated data
print(pii_dataset[['text', 'True Predictions']].head())
print(f"Annotated data successfully written to {annotated_csv_file_path}")

"""# Training Model"""

# Load the dataset
dataset = pd.read_csv(r'Training_Set.csv')

# Convert the dataset into the required format
training_data = []
for index, row in dataset.iterrows():
    text = row['text']
    # Assuming 'True Predictions' is the column containing the annotations
    entities = ast.literal_eval(row['True Predictions'])
    training_data.append((text, {"entities": entities}))

# Display a sample of the training data
print(training_data[:2])

warnings.filterwarnings("ignore", category=UserWarning, module="spacy.training.iob_utils")

# Define the function to merge overlapping entities
def merge_overlapping_entities(entities):
    if not entities:
        return []
    # Sort entities by their start positions
    entities = sorted(entities, key=lambda x: x[0])
    merged_entities = []
    current_start, current_end, current_label = entities[0]

    for start, end, label in entities[1:]:
        if start <= current_end:  # Overlapping
            current_end = max(current_end, end)
        else:
            merged_entities.append((current_start, current_end, current_label))
            current_start, current_end, current_label = start, end, label
    merged_entities.append((current_start, current_end, current_label))
    return merged_entities

# Initialize the Custom Spacy model
nlp = spacy.blank("en")

# Create an NER component and add it to the pipeline
ner = nlp.add_pipe("ner")

# Prepare the training data with merged entities
training_data = []
for index, row in dataset.iterrows():
    text = row['text']
    entities = ast.literal_eval(row['True Predictions'])
    entities = merge_overlapping_entities(entities)

    # Check alignment
    try:
        tags = offsets_to_biluo_tags(nlp.make_doc(text), entities)
        training_data.append((text, {"entities": entities}))
    except Exception as e:
        print(f"Skipping misaligned entity in record {index}: {e}")

# Add the labels to the NER component
for _, annotations in training_data:
    for ent in annotations.get("entities"):
        ner.add_label(ent[2])

# Convert training data to Spacy's format
examples = []
for text, annotations in training_data:
    doc = nlp.make_doc(text)
    example = Example.from_dict(doc, annotations)
    examples.append(example)

# Initialize the optimizer
optimizer = nlp.begin_training()

# Parameters
iterations = 20  # Number of iterations
dropout = 0.5  # Dropout rate
batch_size_start = 4  # Start of the batch size range
batch_size_end = 32  # End of the batch size range

# Training loop
for i in range(iterations):
    losses = {}
    batches = minibatch(examples, size=compounding(batch_size_start, batch_size_end, 1.001))
    for batch in batches:
        nlp.update(batch, losses=losses, drop=dropout, sgd=optimizer)
    print(f"Iteration {i + 1}, Losses: {losses}")

# Save the trained model
output_dir = r'PII Model'
nlp.to_disk(output_dir)
print(f"Model saved to {output_dir}")

"""# Test Dataset Generation"""

# Initialize Faker
fake = Faker()

# Function to generate phone numbers in a specific format
def generate_phone_number():
    formats = [
        '+91 ##########',
        '+## ##########',
        '+### ##########'
    ]
    format_choice = fake.random.choice(formats)
    return fake.numerify(format_choice)

# Function to generate synthetic PII data
def generate_pii_data(num_samples):
    data = {
        "name": [fake.name() for _ in range(num_samples)],
        "credit_card": [fake.credit_card_full() for _ in range(num_samples)],
        "email": [fake.email() for _ in range(num_samples)],
        "url": [fake.url() for _ in range(num_samples)],
        "phone": [generate_phone_number() for _ in range(num_samples)],
        "address": [fake.address() for _ in range(num_samples)],
        "company": [fake.company() for _ in range(num_samples)],
        "ssn": [fake.ssn() for _ in range(num_samples)]
    }
    return pd.DataFrame(data)

# Function to randomly remove full stops
def remove_random_full_stops(text, removal_probability=0.3):
    if random.random() < removal_probability:
        text = text.replace('.', '', random.randint(1, text.count('.')))
    return text

# Generate Dataset with any Number of Samples
pii_dataset = generate_pii_data(100)  

# Combined sentence/audit templates with PII data
sentence_templates = [

    ('''We have conducted a thorough review of the tax compliance practices followed by {company} for the fiscal year ending March 31, 2023. Our examination included a detailed analysis of corporate tax returns, GST filings, and withholding tax submissions across all divisions. The review focused on ensuring compliance with the latest amendments in tax laws and regulations.

    Corporate Tax Overview
    The corporate tax computation for {company} was cross-verified against the financial statements audited by our internal team. The tax liability was calculated considering various deductions under section 80C, 80D, and other relevant sections of the Income Tax Act. The total taxable income stood at INR 500 Crores, with an effective tax rate of 25%.

    Key points include:

    Depreciation Deductions: Claimed as per the Income Tax Act, aligned with the rates prescribed under Schedule II. The assets located at {address} were correctly depreciated using the Written Down Value (WDV) method. The details of high-value assets have been corroborated with the asset register maintained at the corporate office.

    Tax Credits: The company has utilized carry-forward losses from previous financial years to offset the current tax liability, reducing the net payable tax. The adjusted tax liability has been duly filed with the tax authorities.

    Deductions: The deductions for contributions to the Employee Provident Fund (EPF) and Gratuity are in compliance with sections 80C and 80D. However, we noted a delay in the deposit of EPF contributions for some employees whose SSNs {ssn} end with ‘4567’. This delay has been flagged, and a provision for potential interest and penalties has been recommended.

    GST Compliance
    The Goods and Services Tax (GST) compliance was reviewed in detail:

    GST Payments: All GST payments were made on time except for minor discrepancies in the month of July. The shortfall of INR 2 Lakhs in GST payments for {company}’s manufacturing unit at {address} was rectified in subsequent months, with interest computed at 18% p.a.

    Input Tax Credit (ITC): ITC claims were verified against the purchase invoices. The ITC related to capital goods purchased by the Research & Development (R&D) division were adequately accounted for. However, it was observed that certain invoices, particularly from vendors identified by the URL {url}, were not uploaded on the GST portal within the stipulated time, leading to an ITC reversal.

    Reconciliation: A reconciliation of GSTR-3B with GSTR-2A was performed, revealing minor mismatches which have been communicated to the concerned department. The finance team has been instructed to follow up with vendors whose SSNs {ssn} match records ending in '7890' to ensure timely filing.

    Withholding Tax (TDS)
    Withholding tax (TDS) was analyzed across various payments made during the year:

    Salaries: TDS on salaries was deducted as per Section 192 of the Income Tax Act. Employee details, including SSNs {ssn}, were cross-checked with the HR records. A mismatch was found in the TDS calculations for employees whose SSNs end in ‘1234’ due to incorrect consideration of their investment declarations.

    Professional Fees: TDS under Section 194J was reviewed, with a specific focus on payments exceeding INR 30,000. One such transaction involving a payment to {name} was identified where TDS was not deducted. The legal team has been notified, and a rectification process has been initiated.

    Rent Payments: TDS on rent payments was calculated correctly, but it was observed that rent agreements for premises at {address} lacked proper documentation. The agreements are currently being reviewed to ensure compliance.

    Audit Recommendations
    Based on our review, we recommend the following actions to mitigate tax risks:

    Timely Payment of Taxes: Ensure that all taxes, including GST and TDS, are paid within the due dates to avoid interest and penalties. The finance team should regularly review the payment schedules, particularly for transactions involving large sums.

    Documentation: Improve the documentation process, especially for transactions involving high-value assets and payments. Ensure that all contracts and agreements are updated and compliant with tax regulations.

    Employee Training: Conduct training sessions for the finance and HR teams on the latest tax amendments and compliance requirements. Emphasize the importance of accurate TDS calculations and timely tax payments.

    Automation of Processes: Consider implementing tax compliance software to automate GST reconciliation, TDS computation, and other tax-related processes. This will reduce manual errors and ensure adherence to compliance timelines.

    Conclusion
    We have attached a detailed report with the findings and recommendations. The finance team at {company} should review this report and initiate the necessary actions. For any queries or further clarifications, please contact {name} at {email} or {phone}. All physical correspondence can be directed to our office at {address}.

    This review reflects our commitment to ensuring that {company} remains compliant with all tax regulations. We appreciate your cooperation during the audit process and look forward to your prompt action on the recommendations.

    '''),

    "Dear {name}, your recent transaction with {company} using credit card {credit_card} was successful. Please contact us at {phone} if you have any questions.",

    ('''To The Members of {company},
    Report on the audit of the Standalone Financial Statements
    Key Audit Matters Auditors’ response to Key Audit Matters
    Property, Plant & Equipment and Intangible Assets
    There are areas where management judgement impacts the carrying value of property, plant and equipment, intangible assets and their respective depreciation/amortisation rates. These include the decision to capitalize or expense costs; the annual asset life review; the timeliness of the capitalization of assets and the use of management assumptions and estimates for the determination or the measurement and recognition criteria for assets retired from active use. Due to the materiality in the context of the Balance Sheet of the Company and the level of judgement and estimates required, we consider this to be an area of significance.
    We assessed the controls in place over the fixed asset cycle, evaluated the appropriateness of the capitalization process, performed tests of details on costs capitalized, the timeliness of the capitalization of the assets and the de-recognition criteria for assets retired from active use. In performing these procedures, we reviewed the judgements made by management including the nature of underlying costs capitalized; determination of realizable value of the assets retired from active use; the appropriateness of asset lives applied in the calculation of depreciation/amortization; the useful lives of assets prescribed in Schedule II to the Act and the useful lives of certain assets as per the technical assessment of the management. We observed that the management has regularly reviewed the aforesaid judgements and there are no material changes.
    Opinion
    We have audited the accompanying standalone financial statements of {company}, which comprise the Balance Sheet as at March 31, 2023, the Statement of Profit and Loss (including Other Comprehensive Income), the Statement of Changes in Equity and the Statement of Cash Flows for the year then ended and notes to the standalone financial statements including a summary of significant accounting policies and other explanatory information in which are incorporated the financial statements for the year ended on that date audited by the Branch Auditors of the Company’s one Branch, namely Research & Development (R&D) division situated at {address}. In our opinion and to the best of our information and according to the explanations given to us, the aforesaid standalone financial statements give the information required by the Companies Act, 2013 (the “Act”) in the manner so required and give a true and fair view in conformity with Indian Accounting Standards specified under section 133 of the Act read with the Companies (Indian Accounting Standards) Rules 2015, as amended and other accounting principles generally accepted in India, of the state of affairs of the Company as at March 31, 2023, and total comprehensive income (comprising of profit and other comprehensive income), changes in equity and its cash flows for the year ended on that date.
    Basis for opinion
    We conducted our audit of the standalone financial statements in accordance with the Standards on Auditing (SAs) specified under section 143(10) of the Act. Our responsibilities under those Standards are further described in the Auditors’ Responsibilities for the Audit of the standalone financial statements section of our report. We are independent of the Company in accordance with the Code of Ethics issued by the Institute of Chartered Accountants of India (“ICAI”), together with the ethical requirements that are relevant to our audit of the standalone financial statements under the provisions of the Act and "the Rules" thereunder, and we have fulfilled our other ethical responsibilities in accordance with these requirements and the ICAI’s Code of Ethics. We believe that the audit evidence we have obtained is sufficient and appropriate to provide a basis for our opinion on the standalone financial statements.
    Key audit matters
    Key audit matters are those matters that, in our professional judgment, were of most significance in our audit of the standalone financial statements of the current period. These matters were addressed in the context of our audit of the standalone financial statements as a whole, and in forming our opinion thereon, and we do not provide a separate opinion on these matters. We have determined the matters described below to be the key audit matters to be communicated in our report.'''),

    ('''Following the annual compliance schedule, a thorough evaluation of the internal controls within {company} was conducted. This document outlines the methodologies adopted, key findings, and recommendations for strengthening internal processes and ensuring regulatory compliance. The objective of this review was to assess the effectiveness of the existing controls and identify potential areas for enhancement.
    The review process included both quantitative and qualitative assessments, leveraging advanced data analytics tools and direct observation techniques. The methodology encompassed:
    Risk Assessment:
    Identification and prioritization of risk areas within the financial and operational domains
    Deployment of risk management frameworks to evaluate the potential impact and likelihood of identified risks
    Control Testing:
    Execution of control tests to verify the functionality and effectiveness of control measures
    Analysis of control gaps and deficiencies in critical areas
    Stakeholder Interviews:
    Conducting interviews with key personnel to gather insights on control environments and operational challenges
    Evaluation of the awareness and understanding of control policies among staff
    Several critical findings emerged from the review, indicating areas that require immediate attention and corrective actions:
    Financial Control Deficiencies:
    Inconsistent application of accounting policies leading to discrepancies in financial reporting
    Lack of adequate documentation for significant financial transactions
    Operational Control Gaps:
    Inefficiencies in the procurement process resulting in unauthorized purchases
    Insufficient monitoring of inventory levels causing stock variances
    Compliance Shortcomings:
    Non-compliance with internal audit recommendations from previous assessments
    Delays in regulatory filings and updates
    To address the identified deficiencies and enhance the control environment, the following actions are recommended:
    Financial Controls:
    Standardization of accounting procedures across all departments
    Complementation of a centralized documentation system for financial transactions
    Operational Controls:
    Revision of the procurement policy to include stricter approval processes
    Regular inventory audits to ensure accuracy and accountability
    Compliance Enhancements:
    Establishment of a compliance oversight committee to monitor adherence to audit recommendations
    Timely updating and submission of regulatory documents
    The review highlights the necessity for continuous improvement in internal controls to mitigate risks and ensure compliance. The implementation of the recommended actions will significantly enhance the operational efficiency and financial integrity of {company}
    Attached to this document are detailed reports and evidence supporting the findings and recommendations. The Internal Compliance Unit is available for further discussions and clarifications
    For any questions or clarifications regarding this document, please contact the Internal Compliance Unit at {email} or {phone}. Physical correspondence can be directed to {address}. Please also include the last four digits of your SSN: {ssn}.
    This document contains proprietary information of {company}. Unauthorized use or disclosure of the contents is strictly prohibited. All related communications should be directed to authorized personnel only.
    Please acknowledge receipt of this document by providing the last four digits of your {credit_card}, your full name {name}, and the associated {url} to our secure email. This step is crucial for maintaining the security and confidentiality of our internal review process.'''),

    "The report for {company} located at {address} was submitted successfully. For further inquiries, reach out at {email}.",

    "The SSN {ssn} associated with your account has been verified. Visit {url} for more details.",

    ('''To Whom It May Concern,

    This letter is to confirm that {name}, holding Social Security Number (SSN) {ssn}, residing at {address}, has filed their tax returns for the fiscal year ending March 31, 2023. The tax filings have been processed under the IRS Tax Identification Number (TIN) associated with the company {company}.

    The individual’s total income for the fiscal year amounted to $125,000, including salaries, bonuses, and other forms of income. The detailed breakdown of the income sources is as follows:

    1. Salary from {company}: $100,000
    2. Bonus and Incentives: $15,000
    3. Other Income (Investments, Dividends, etc.): $10,000

    The total federal tax liability for the year is $25,000, which has been fully paid by the taxpayer. The tax payments were made using the credit card ending in {credit_card} and were processed through the IRS online payment portal. Please note that this tax statement is generated in accordance with the income tax laws applicable in the United States.

    For any queries or further clarifications, you may contact {name} at {email} or {phone}. Additional documents supporting the tax filings can be requested by visiting {url}.

    This document is confidential and should be handled in accordance with data protection regulations to prevent unauthorized access to sensitive information.
    '''),

    ('''Subject: Annual Tax Filing Confirmation for Fiscal Year 2023

    Dear {name},

    We are pleased to inform you that your tax return for the fiscal year ending March 31, 2023, has been successfully filed and processed by {company}. The filing was completed using your Social Security Number (SSN) {ssn}, and the confirmation number is associated with the TIN {ssn} registered under {company}.

    Your gross income for the year was reported as $150,000, which includes:

    - Employment Income from {company}: $120,000
    - Capital Gains: $20,000
    - Interest and Dividends: $10,000

    The total tax due for the fiscal year was calculated at $30,000. This amount has been paid in full through a payment transaction completed on March 28, 2024, using the credit card ending in {credit_card}. Your tax records indicate that you are eligible for a tax refund of $2,000, which will be credited to your bank account on file.

    Please ensure that all records related to this tax filing, including the payment receipt, are stored securely. Should you require any further assistance or have questions about your tax return, please do not hesitate to contact our customer service department at {email} or by calling {phone}. You may also visit our website {url} for more information.

    This statement is intended for the use of {name} and contains sensitive information that must be kept confidential. Any unauthorized use, dissemination, or copying of this document is strictly prohibited.
    '''),

    ('''To: {name}
    SSN: {ssn}
    TIN: {ssn}
    Address: {address}

    Subject: Confirmation of Tax Filing for FY 2023

    Dear {name},

    Your tax filing for the fiscal year ending March 31, 2023, has been successfully processed. The filing was conducted using your TIN {ssn} registered under the IRS. Your total income for the fiscal year was reported as $175,000, comprising the following sources:

    1. Salary from {company}: $140,000
    2. Investment Income: $25,000
    3. Other Earnings: $10,000

    The total tax payable for this fiscal year was calculated at $35,000. The payment was made on March 25, 2024, using the credit card associated with the number ending in {credit_card}. Please retain this statement as proof of payment and tax compliance.

    Should you need to amend any details or have inquiries, you can contact us at {email} or by phone at {phone}. Additional information and related services can be accessed through {url}.

    Please note that this document contains confidential information, and unauthorized access or distribution is strictly prohibited. Keep this document in a secure location.
    '''),

    "During the audit of the financial statements for {company}, it was observed that Mr. {name}, the Chief Financial Officer, approved the purchase of assets worth $500,000 on {address}. The payment was processed through the credit card ending in {credit_card}. Mr. {name}'s Social Security Number (SSN) is {ssn}. For any clarifications, please reach out to Mr. {name} at {email} or contact him directly at {phone}. Further details can be accessed at {url}.",

    "Customer {name} from {company}, located at {address}, has submitted feedback regarding the recent transaction involving their credit card {credit_card}. They can be reached at {email} or {phone} for further discussions. The customer's Social Security Number (SSN) on file is {ssn}. The feedback was originally submitted through our website at {url}.",

    "This is a confirmation that the payment for the invoice number INV-409876535422 from {company} has been successfully processed. The payment was made using the credit card ending in {credit_card} by {name}. The billing address on file is {address}. The Social Security Number (SSN) for {name} is {ssn}. Should you have any inquiries, you may contact {name} via email at {email} or call {phone}. For more information, visit {url}",

    "We regret to inform you that a security breach was detected on {company}'s systems, which may have exposed your personal information, including your name ({name}), email ({email}), phone number ({phone}), and Social Security Number (SSN) ({ssn}). The breach was traced back to unauthorized access from IP address 192.0.45. If you notice any suspicious activity on your credit card ending in {credit_card}, please contact us immediately. You can also check for updates on our security measures at {url}. The compromised data was stored at our facility located at {address}.",

    "Dear {name}, thank you for creating a new account with {company}. Your registered email is {email}, your contact number is {phone}, and your Social Security Number (SSN) is {ssn}. The account was set up using the billing address {address}, and the primary credit card linked to the account ends in {credit_card}. Please visit {url} to verify your account and update any personal details. If you need assistance, contact our support team",

    "This Service Contract between {company} and {name} was entered at {address}. The contract stipulates that all payments will be processed through the credit card provided by {name}, ending in {credit_card}. {name}'s Social Security Number (SSN) is {ssn}. For further reference, correspondence will be sent to {email}, and all communications will be conducted via {phone}. The full contract details are available online at {url}.",

    "Dear {name}, we have received your loan application at {company}, and it is currently under review. Your application, submitted on 23/09/2008, includes personal details such as your home address ({address}), email ({email}), contact number ({phone}), and Social Security Number (SSN) ({ssn}). The loan amount requested will be credited to your account associated with the credit card ending in {credit_card}. Please check {url} for real-time updates on your application status.",

    "Insurance claim #CLM9076109877 has been initiated by {name} for {company}. The claim, associated with the address {address}, will be processed through the credit card ending in {credit_card}. Our claims department may reach out to you at {phone} or via email at {email} for additional information. {name}'s Social Security Number (SSN) is {ssn}. Claim details are available online at {url}.",

    "We are pleased to welcome {name} to {company}. As part of the onboarding process, we have registered your personal details, including your residential address ({address}), contact number ({phone}), email ({email}), and Social Security Number (SSN) ({ssn}). Your corporate credit card, ending in {credit_card}, will be issued within the next five business days. For company policies and other relevant information, please visit {url}.",

    "Dear {name}, your subscription with {company} is up for renewal. The subscription associated with the email {email}, phone number {phone}, and billing address {address} will automatically renew using your credit card ending in {credit_card}. Your Social Security Number (SSN) on file is {ssn}. To manage your subscription or for more details, please visit {url}. If you need to update your payment information, contact our support team."

    "During the routine financial audit for {company}, it was discovered that Mr. {name}, serving as the CFO, authorized the acquisition of assets totaling $500,000. The transaction occurred at {address} and was paid for using the credit card ending in {credit_card}. Mr. {name} is identified by the Social Security Number (SSN) {ssn}. For additional details, Mr. {name} can be contacted at {email} or via phone at {phone}. Further documentation can be found on our website at {url}",

    "Customer feedback has been received from {name} representing {company}, located at {address}. The feedback pertains to a transaction processed through their credit card ending in {credit_card}. You can reach out to {name} for more information at {email} or {phone}. The SSN associated with the customer's profile is {ssn}. The feedback submission was completed through our online portal at {url}.",

    "We hereby confirm that the payment for invoice INV-12345 issued by {company} has been successfully completed. The payment was facilitated by {name} using the credit card ending in {credit_card}. The billing address linked to the payment is {address}. The associated SSN for {name} is {ssn}. For inquiries, please contact {name} at {email} or by phone at {phone}. Visit {url} for further details",

    "A security incident has been detected on the systems of {company}, potentially compromising your personal data, including your name ({name}), email ({email}), phone number ({phone}), and SSN ({ssn}). The breach was linked to an unauthorized access attempt traced to IP address. If you observe any suspicious activity on your credit card ending in {credit_card}, notify us immediately. Updates on the situation will be posted at {url}. The compromised data was stored at our facility located at {address}",

    "Dear {name}, we are pleased to confirm the creation of your new account with {company}. Your registered email address is {email}, and your contact number is {phone}. The account was set up using your billing address {address} and is linked to a credit card ending in {credit_card}. Your Social Security Number (SSN) is {ssn}. Please visit {url} to verify your account and update any personal information. If you require assistance, our support team is available to help.",

    "This agreement between {company} and {name} was formalized at {address}. Under the terms of the contract, all payments will be processed via the credit card provided by {name}, ending in {credit_card}. {name}'s Social Security Number (SSN) is {ssn}. All correspondence will be directed to {email}, and further communication can be made via {phone}. Full contract details are accessible online at {url}",

    "Dear {name}, your recent loan application at {company} is currently under review. The application, submitted on 09/12/2020, includes your home address ({address}), email ({email}), phone number ({phone}), and SSN ({ssn}). The requested loan amount will be credited to the account linked to the credit card ending in {credit_card}. Check {url} for real-time updates on your application status",

    "Insurance claim #CLM0098 has been initiated by {name} with {company}. The claim, related to the address {address}, will be processed using the credit card ending in {credit_card}. Our claims department may contact you at {phone} or {email} for further details. The SSN associated with {name} is {ssn}. Claim details can be accessed at {url}.",

    "We are excited to welcome {name} to {company}. As part of your onboarding, we have registered your personal details, including your home address ({address}), contact number ({phone}), email ({email}), and SSN ({ssn}). Your corporate credit card, ending in {credit_card}, will be issued shortly. For more information on company policies, please visit {url}.",

    "Dear {name}, your subscription with {company} is approaching its renewal date. The subscription linked to the email {email}, phone number {phone}, and billing address {address} will automatically renew using the credit card ending in {credit_card}. The SSN on file for this account is {ssn}. To manage your subscription or update payment details, visit {url}. If you need further assistance, please contact our support team.",

    "During the audit of the financial statements for {company}, it was observed that Mr. {name}, the Chief Financial Officer, approved the purchase of assets worth $500,000 on {address}. The payment was processed through the credit card ending in {credit_card}. Mr. {name}'s Social Security Number (SSN) is {ssn}. For any clarifications, please reach out to Mr. {name} at {email} or contact him directly at {phone}. Further details can be accessed at {url}.",

    "Customer {name} from {company}, located at {address}, has submitted feedback regarding the recent transaction involving their credit card {credit_card}. They can be reached at {email} or {phone} for further discussions. The customer's Social Security Number (SSN) on file is {ssn}. The feedback was originally submitted through our website at {url}.",

    "This is a confirmation that the payment for the invoice number INV-3066 from {company} has been successfully processed. The payment was made using the credit card ending in {credit_card} by {name}. The billing address on file is {address}. The Social Security Number (SSN) for {name} is {ssn}. Should you have any inquiries, you may contact {name} via email at {email} or call {phone}. For more information, visit {url}.",

    "We regret to inform you that a security breach was detected on {company}'s systems, which may have exposed your personal information, including your name ({name}), email ({email}), phone number ({phone}), and Social Security Number (SSN) ({ssn}). The breach was traced back to unauthorized access from IP address 190.22.99. If you notice any suspicious activity on your credit card ending in {credit_card}, please contact us immediately. You can also check for updates on our security measures at {url}. The compromised data was stored at our facility located at {address}.",

]

# Apply the templates to generate sentences with PII data
pii_dataset['text'] = pii_dataset.apply(lambda row: sentence_templates[row.name % len(sentence_templates)].format(
    name=row['name'],
    company=row['company'],
    email=row['email'],
    url=row['url'],
    phone=row['phone'],
    address=row['address'],
    credit_card=row['credit_card'],
    ssn=row['ssn']
), axis=1)

# Apply the random full stop removal
pii_dataset['text'] = pii_dataset['text'].apply(remove_random_full_stops)

# Annotate the PII data in text
def annotate_pii(text, pii_dict):
    annotations = []
    for pii_type, pii_value in pii_dict.items():
        escaped_pii_value = re.escape(pii_value)
        matches = list(re.finditer(escaped_pii_value, text))
        for match in matches:
            start, end = match.span()
            annotations.append((start, end, pii_type))
    return annotations

# Apply the annotation function to each row
pii_dataset['True Predictions'] = pii_dataset.apply(lambda row: annotate_pii(
    row['text'], {
        'name': row['name'],
        'credit_card': row['credit_card'],
        'email': row['email'],
        'url': row['url'],
        'phone': row['phone'],
        'address': row['address'],
        'company': row['company'],
        'ssn': row['ssn']
    }), axis=1)

# Save the testing dataset to a CSV file
testing_csv_file_path = r'Testing_Set.csv'
pii_dataset.to_csv(testing_csv_file_path, index=False)

# Display the first few rows of the testing dataset
print(pii_dataset[['text', 'True Predictions']].head())
print(f"Testing data successfully written to {testing_csv_file_path}")

"""# Testing Code to get Results"""

# Load the trained model
model_dir = r"PII Model"
nlp = spacy.load(model_dir)

# Load the test dataset
test_dataset_path = r"Testing_Set.csv"
test_dataset = pd.read_csv(test_dataset_path)

# Function to format the predictions from the model
def get_model_predictions(text):
    doc = nlp(text)
    predictions = []

    # Initialize dictionary to collect predicted fields
    predicted_entities = {
        'Name': [],
        'Email': [],
        'Url': [],
        'Phone': [],
        'Address': [],
        'Company': [],
        'Credit_card': [],
        'SSN': []
    }

    for ent in doc.ents:
        entity_label = ent.label_.lower()  # Keep the label lowercase for the predictions list
        predictions.append((ent.start_char, ent.end_char, entity_label))

        # Map lowercase labels to capitalized column names
        if entity_label == 'name':
            predicted_entities['Name'].append(ent.text)
        elif entity_label == 'email':
            predicted_entities['Email'].append(ent.text)
        elif entity_label == 'url':
            predicted_entities['Url'].append(ent.text)
        elif entity_label == 'phone':
            predicted_entities['Phone'].append(ent.text)
        elif entity_label == 'address':
            predicted_entities['Address'].append(ent.text)
        elif entity_label == 'company':
            predicted_entities['Company'].append(ent.text)
        elif entity_label == 'credit_card':
            predicted_entities['Credit_card'].append(ent.text)
        elif entity_label == 'ssn':
            predicted_entities['SSN'].append(ent.text)

    return predictions, predicted_entities

# Function to compare the true annotations with model predictions
def evaluate_predictions(true_annotations, predicted_annotations):
    true_annotations_set = set(true_annotations)
    predicted_annotations_set = set(predicted_annotations)

    # Calculate true positives, false positives, and false negatives
    true_positives = len(true_annotations_set & predicted_annotations_set)
    false_positives = len(predicted_annotations_set - true_annotations_set)
    false_negatives = len(true_annotations_set - predicted_annotations_set)

    return true_positives, false_positives, false_negatives

# Lists to store overall true and predicted values for metrics calculation
overall_true_annotations = []
overall_predicted_annotations = []

# Iterate through the test dataset and evaluate the model
for index, row in test_dataset.iterrows():
    text = row['text']
    true_annotations = ast.literal_eval(row['True Predictions'])
    predicted_annotations, predicted_entities = get_model_predictions(text)

    true_positives, false_positives, false_negatives = evaluate_predictions(true_annotations, predicted_annotations)

    # Store the predictions in the DataFrame
    test_dataset.at[index, 'Predicted Results'] = str(predicted_annotations)
    for key, value in predicted_entities.items():
        if value:  # Only add non-empty lists
            test_dataset.at[index, 'Predicted ' + key] = ', '.join(value)

    # Extend the overall lists for metric calculation
    overall_true_annotations.extend([1] * len(true_annotations))
    overall_true_annotations.extend([0] * (len(predicted_annotations) - len(true_annotations)))

    overall_predicted_annotations.extend([1] * len(predicted_annotations))
    overall_predicted_annotations.extend([0] * (len(true_annotations) - len(predicted_annotations)))

# Calculate precision, recall, F1 score, and accuracy
precision = precision_score(overall_true_annotations, overall_predicted_annotations)
recall = recall_score(overall_true_annotations, overall_predicted_annotations)
f1 = f1_score(overall_true_annotations, overall_predicted_annotations)
accuracy = accuracy_score(overall_true_annotations, overall_predicted_annotations)

# Save the test dataset with predictions to a new CSV file
output_test_dataset_path = r'Results.xlsx'
test_dataset.to_excel(output_test_dataset_path, index=False)

print(f"Predictions saved to {output_test_dataset_path}")

"""# Anonymization of Texts to Anonymize PII Data in Texts"""

# Load the test results Excel file
results_df = pd.read_excel(r"Results.xlsx")

# Rename 'True Predictions' to 'True Results'
results_df.rename(columns={'True Predictions': 'True Results'}, inplace=True)

# Create a new Excel workbook
workbook = Workbook()

# Access the active worksheet and rename it to 'Audit Reports'
audit_reports_sheet = workbook.active
audit_reports_sheet.title = 'Audit Reports'

# 1. Create the 'Audit Reports' section
# Copy the 'text' column from the results_df DataFrame and rename it to 'Text'
audit_reports_df = results_df[['text']].copy()
audit_reports_df.rename(columns={'text': 'Text'}, inplace=True)

# Write the 'Audit Reports' section to the Excel file
for r in dataframe_to_rows(audit_reports_df, index=False, header=True):
    audit_reports_sheet.append(r)

# 2. Create the 'Predicted Results' section
# Create a new sheet for 'Predicted Results'
predicted_results_sheet = workbook.create_sheet(title='Predicted Results')

# Rearrange the columns in the specified order
predicted_results_df = results_df[[
    'True Results', 'Predicted Results', 'Predicted Name', 'Predicted Phone',
    'Predicted Email', 'Predicted Address', 'Predicted SSN',
    'Predicted Credit_card', 'Predicted Company', 'Predicted Url'
]]

# Rename the columns to the format you requested
predicted_results_df.columns = [
    'True Results', 'Predicted Results', 'Name', 'Phone Number',
    'Email', 'Address', 'SSN',
    'Credit Card', 'Company Name', 'URL'
]

# Write the 'Predicted Results' section to the Excel file
for r in dataframe_to_rows(predicted_results_df, index=False, header=True):
    predicted_results_sheet.append(r)

# 3. Anonymize the data based on the Predicted Results
def anonymize_text(text, predictions):
    replacements = {
        'name': '[NAME REDACTED]',
        'email': '[EMAIL REDACTED]',
        'url': '[URL REDACTED]',
        'phone': '[PHONE NUMBER REDACTED]',
        'address': '[ADDRESS REDACTED]',
        'company': '[COMPANY NAME REDACTED]',
        'credit_card': '[CREDIT CARD REDACTED]',
        'ssn': '[SSN REDACTED]'
    }

    # Sort predictions by the start position to avoid replacing wrong indices after text manipulation
    predictions = sorted(predictions, key=lambda x: x[0], reverse=True)

    for start, end, label in predictions:
        pii_text = text[start:end]
        if label in replacements:
            text = text[:start] + replacements[label] + text[end:]
    return text

# Create a new DataFrame for the anonymized data
anonymized_data_df = results_df[['text', 'Predicted Results']].copy()
anonymized_data_df.rename(columns={'text': 'Original Text'}, inplace=True)

# Apply the anonymization
anonymized_data_df['Anonymized Text'] = anonymized_data_df.apply(
    lambda row: anonymize_text(row['Original Text'], ast.literal_eval(row['Predicted Results'])), axis=1
)

# 4. Create the 'Anonymized Data' section
# Create a new sheet for 'Anonymized Data'
anonymized_data_sheet = workbook.create_sheet(title='Anonymized Data')

# Write the 'Anonymized Data' section to the Excel file
for r in dataframe_to_rows(anonymized_data_df, index=False, header=True):
    anonymized_data_sheet.append(r)

# Save the workbook to a file
output_file = r'Results.xlsx'
workbook.save(output_file)

print(f"Results formatted and Anonymized and saved to {output_file}")

"""# Generating Graphs and Matrix"""

# Load the Excel file with multiple sheets
file_path = r"Results.xlsx"  
xls = pd.ExcelFile(file_path)

# Load the "Predicted Results" sheet
predicted_results_df = pd.read_excel(xls, sheet_name='Predicted Results')

# Function to parse the string representations of entities
def parse_entities(entity_string):
    if isinstance(entity_string, str):
        return eval(entity_string)
    return []

# Parse the True and Predicted Results
predicted_results_df['True Parsed'] = predicted_results_df['True Results'].apply(parse_entities)
predicted_results_df['Predicted Parsed'] = predicted_results_df['Predicted Results'].apply(parse_entities)

# Initialize lists for confusion matrix
true_labels = []
predicted_labels = []
predicted_probabilities = []

# Iterate through each row to compare and prepare confusion matrix data
for idx, row in predicted_results_df.iterrows():
    true_entities = {(start, end, label) for start, end, label in row['True Parsed']}
    predicted_entities = {(start, end, label) for start, end, label in row['Predicted Parsed']}

    # Align true and predicted entities by their labels
    for entity in true_entities:
        start, end, label = entity
        if entity in predicted_entities:
            true_labels.append(label)
            predicted_labels.append(label)
            predicted_probabilities.append(1)  # Assuming perfect prediction for demonstration
        else:
            true_labels.append(label)
            predicted_labels.append('O')  # 'O' for incorrect prediction or missed detection
            predicted_probabilities.append(0)  # Assuming missed prediction

    for entity in predicted_entities:
        if entity not in true_entities:
            _, _, label = entity
            true_labels.append('O')
            predicted_labels.append(label)
            predicted_probabilities.append(0.5)  # Assuming some confidence for incorrect prediction

# Calculate average metrics
precision, recall, f1, _ = precision_recall_fscore_support(true_labels, predicted_labels, average='weighted')
accuracy = accuracy_score(true_labels, predicted_labels)

# Print the average metrics
print(f"Precision: {precision}")
print(f"Recall: {recall}")
print(f"F1 Score: {f1}")
print(f"Accuracy: {accuracy}")

# Data for plotting
metrics = ['Precision', 'Recall', 'F1 Score', 'Accuracy']
values = [precision, recall, f1, accuracy]

# Plotting the bar chart
plt.figure(figsize=(10, 6))
plt.bar(metrics, values, color=['skyblue', 'orange', 'green', 'red'])
plt.ylim(0, 1)
plt.title('Model Performance Metrics')
plt.xlabel('Metrics')
plt.ylabel('Score')
plt.show()

# Generate the confusion matrix
labels = sorted(set(true_labels + predicted_labels))  # Get unique labels
conf_matrix = confusion_matrix(true_labels, predicted_labels, labels=labels)

# Plot the confusion matrix heatmap
plt.figure(figsize=(12, 8))
sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues', xticklabels=labels, yticklabels=labels)
plt.title('Confusion Matrix for PII Detection')
plt.xlabel('Predicted Labels')
plt.ylabel('True Labels')
plt.show()

# ROC Curve for each class
plt.figure(figsize=(10, 8))
for i, label in enumerate(labels):
    y_true = [1 if l == label else 0 for l in true_labels]
    y_pred = [1 if l == label else 0 for l in predicted_labels]
    fpr, tpr, _ = roc_curve(y_true, y_pred)
    auc_score = roc_auc_score(y_true, y_pred)
    plt.plot(fpr, tpr, label=f'{label} (AUC = {auc_score:.2f})')

plt.plot([0, 1], [0, 1], color='navy', linestyle='--')  # Diagonal line for random classifier
plt.title('ROC Curve')
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.legend(loc='lower right')
plt.show()

# Precision-Recall curve for each class
plt.figure(figsize=(10, 8))
for i, label in enumerate(labels):
    y_true = [1 if l == label else 0 for l in true_labels]
    y_pred = [1 if l == label else 0 for l in predicted_labels]
    precision, recall, _ = precision_recall_curve(y_true, y_pred)
    plt.plot(recall, precision, label=label)

plt.title('Precision-Recall Curve')
plt.xlabel('Recall')
plt.ylabel('Precision')
plt.legend()
plt.show()